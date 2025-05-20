import tornado.ioloop
import tornado.web
import sqlite3
import json
import re
import requests
import settings
import mail_handler
import datetime  # Importiamo una sola volta il modulo datetime
from datetime import date  # Importiamo solo date dalla libreria datetime

# Credenziali fisse per l'accesso
USERNAME = "a"
PASSWORD = "aa"

# ... resto del codice ...

def validate_association_data(data):
    errors = []
    
    if not data.get('id_studente'):
        errors.append("ID studente mancante")
    
    if not data.get('id_azienda'):
        errors.append("ID azienda mancante")
    
    # Validazione date
    data_inizio = data.get('data_inizio')
    data_fine = data.get('data_fine')
    
    # Verifica che le date siano in formato corretto (YYYY-MM-DD)
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    
    if data_inizio and not date_pattern.match(data_inizio):
        errors.append("Formato data inizio non valido, usa YYYY-MM-DD")
    
    if data_fine and not date_pattern.match(data_fine):
        errors.append("Formato data fine non valido, usa YYYY-MM-DD")
    
    # Verifica che la data fine sia successiva alla data inizio
    if data_inizio and data_fine:
        try:
            inizio = datetime.datetime.strptime(data_inizio, '%Y-%m-%d').date()
            fine = datetime.datetime.strptime(data_fine, '%Y-%m-%d').date()
            
            if fine < inizio:
                errors.append("La data di fine deve essere successiva alla data di inizio")
        except ValueError:
            # Gli errori di formato sono già gestiti sopra
            pass
    
    return errors

def validate_student_data(data):
    """
    Validate student data before insertion or update.
    """
    errors = []
    if not data.get('nome'):
        errors.append('Nome è obbligatorio')
    email = data.get('email', '')
    if not email or not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        errors.append('Email non valida')
    # Non facciamo validazione obbligatoria per il campo classe
    # poiché potrebbe essere opzionale
    return errors



def validate_azienda_data(data):
    """
    Validate company data before insertion or update.
    """
    errors = []
    if not data.get('nome'):
        errors.append('Nome azienda è obbligatorio')
    return errors

class AddAssociazioneHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return

        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return

            # Validate input
            validation_errors = validate_association_data(data)
            if validation_errors:
                self.set_status(400)
                self.write({"success": False, "errors": validation_errors})
                return

            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO associazioni (id_studente, id_azienda, data_inizio, data_fine)
                VALUES (?, ?, ?, ?)
            """, (data.get('id_studente'), data.get('id_azienda'), 
                  data.get('data_inizio'), data.get('data_fine')))

            last_id = cursor.lastrowid
            conn.commit()
            conn.close()

            self.write({"success": True, "id": last_id})

        except Exception as e:
            self.set_status(500)
            print(f"Errore durante aggiunta associazione: {str(e)}")
            self.write({"success": False, "error": str(e)})

class UpdateAssociazioneHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return

        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return
                
            # Controlla che l'ID sia presente
            if not data.get('id'):
                self.set_status(400)
                self.write({"success": False, "error": "ID associazione mancante"})
                return

            # Validate input
            validation_errors = validate_association_data(data)
            if validation_errors:
                self.set_status(400)
                self.write({"success": False, "errors": validation_errors})
                return

            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE associazioni
                SET id_studente = ?, id_azienda = ?, data_inizio = ?, data_fine = ?
                WHERE id = ?
            """, (data.get('id_studente'), data.get('id_azienda'), 
                  data.get('data_inizio'), data.get('data_fine'), data.get('id')))

            conn.commit()
            conn.close()

            self.write({"success": True})

        except Exception as e:
            self.set_status(500)
            print(f"Errore durante aggiornamento associazione: {str(e)}")
            self.write({"success": False, "error": str(e)})

class DeleteAssociazioneHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return

        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return
                
            # Controlla che l'ID sia presente
            if not data.get('id'):
                self.set_status(400)
                self.write({"success": False, "error": "ID associazione mancante"})
                return

            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            cursor.execute("DELETE FROM associazioni WHERE id = ?", (data.get('id'),))

            conn.commit()
            conn.close()

            self.write({"success": True})

        except Exception as e:
            self.set_status(500)
            print(f"Errore durante eliminazione associazione: {str(e)}")
            self.write({"success": False, "error": str(e)})

class GetAssociazioniHandler(tornado.web.RequestHandler):     
    def get(self):         
        try:             
            conn = sqlite3.connect("db.sqlite3")             
            cur = conn.cursor()                          
            
            # Verifichiamo prima se esiste il campo classe nella tabella studenti             
            cur.execute("PRAGMA table_info(studenti)")             
            columns = cur.fetchall()             
            column_names = [column[1] for column in columns]             
            classe_exists = 'classe' in column_names                          
            
            # Gestiamo la query in base alla presenza del campo classe             
            if classe_exists:                 
                cur.execute("""                     
                    SELECT a.id, s.nome as studente_nome, s.email as studente_email,                           
                          s.classe as studente_classe,                           
                          az.nome as azienda_nome, az.comune as azienda_comune,                           
                          a.id_studente, a.id_azienda, a.data_inizio, a.data_fine                     
                    FROM associazioni a                     
                    JOIN studenti s ON a.id_studente = s.id                     
                    JOIN aziende az ON a.id_azienda = az.id                 
                """)             
            else:                 
                cur.execute("""                     
                    SELECT a.id, s.nome as studente_nome, s.email as studente_email,                           
                          az.nome as azienda_nome, az.comune as azienda_comune,                           
                          a.id_studente, a.id_azienda, a.data_inizio, a.data_fine                     
                    FROM associazioni a                     
                    JOIN studenti s ON a.id_studente = s.id                     
                    JOIN aziende az ON a.id_azienda = az.id                 
                """)                          
            
            rows = cur.fetchall()             
            data = []                          
            
            for row in rows:                 
                if classe_exists:                     
                    associazione = {                         
                        "id": row[0],                         
                        "studente_nome": row[1] if row[1] is not None else "",                         
                        "studente_email": row[2] if row[2] is not None else "",                         
                        "studente_classe": row[3] if row[3] is not None else "",                         
                        "azienda_nome": row[4] if row[4] is not None else "",                         
                        "azienda_comune": row[5] if row[5] is not None else "",                         
                        "id_studente": row[6] if row[6] is not None else 0,                         
                        "id_azienda": row[7] if row[7] is not None else 0,
                        "data_inizio": row[8] if row[8] is not None else "",
                        "data_fine": row[9] if row[9] is not None else ""
                    }                 
                else:                     
                    associazione = {                         
                        "id": row[0],                         
                        "studente_nome": row[1] if row[1] is not None else "",                         
                        "studente_email": row[2] if row[2] is not None else "",                         
                        "studente_classe": "", # Campo classe vuoto se non esiste nella tabella                         
                        "azienda_nome": row[3] if row[3] is not None else "",                         
                        "azienda_comune": row[4] if row[4] is not None else "",                         
                        "id_studente": row[5] if row[5] is not None else 0,                         
                        "id_azienda": row[6] if row[6] is not None else 0,
                        "data_inizio": row[7] if row[7] is not None else "",
                        "data_fine": row[8] if row[8] is not None else ""
                    }                 
                data.append(associazione)                          
            
            conn.close()             
            self.write({"associazioni": data})         
        except Exception as e:             
            self.set_status(500)             
            print(f"Errore durante recupero associazioni: {str(e)}")             
            self.write({"error": str(e)})

# Funzione di validazione per i dati delle associazioni
def validate_association_data(data):
    errors = []
    
    if not data.get('id_studente'):
        errors.append("ID studente mancante")
    
    if not data.get('id_azienda'):
        errors.append("ID azienda mancante")
    
    # Validazione date
    data_inizio = data.get('data_inizio')
    data_fine = data.get('data_fine')
    
    # Verifica che le date siano in formato corretto (YYYY-MM-DD)
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    
    if data_inizio and not date_pattern.match(data_inizio):
        errors.append("Formato data inizio non valido, usa YYYY-MM-DD")
    
    if data_fine and not date_pattern.match(data_fine):
        errors.append("Formato data fine non valido, usa YYYY-MM-DD")
    
    # Verifica che la data fine sia successiva alla data inizio
    if data_inizio and data_fine:
        try:
            inizio = datetime.datetime.strptime(data_inizio, '%Y-%m-%d').date()
            fine = datetime.datetime.strptime(data_fine, '%Y-%m-%d').date()
            
            if fine < inizio:
                errors.append("La data di fine deve essere successiva alla data di inizio")
        except ValueError:
            # Gli errori di formato sono già gestiti sopra
            pass
    
    return errors
class AddAziendaHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return

        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return
            
            # Validate input
            validation_errors = validate_azienda_data(data)
            if validation_errors:
                self.set_status(400)
                self.write({"success": False, "errors": validation_errors})
                return
            
            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            
            # Verifichiamo prima se esiste la colonna stage
            cursor.execute("PRAGMA table_info(aziende)")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns]
            stage_exists = 'stage' in column_names
            
            # Se non esiste, aggiungiamola
            if not stage_exists:
                cursor.execute("ALTER TABLE aziende ADD COLUMN stage BOOLEAN DEFAULT 0")
                conn.commit()
            
            cursor.execute("""
                INSERT INTO aziende (
                    nome,
                    via,
                    cap,
                    comune,
                    nome_referente,
                    email_referente,
                    cellulare_referente,
                    disponibilita_lavoro_estivo,
                    osservazioni,
                    stage
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data.get("nome", ""),
                data.get("via", ""),
                data.get("cap", ""),
                data.get("comune", ""),
                data.get("nome_referente", ""),
                data.get("email_referente", ""),
                data.get("cellulare_referente", ""),
                data.get("disponibilita_lavoro_estivo", False),
                data.get("osservazioni", ""),
                data.get("stage_pcto", False)  # Modificato per accettare il campo stage_pcto dal frontend
            ))
            last_id = cursor.lastrowid
            conn.commit()
            conn.close()
            self.write({"success": True, "id": last_id})
        except Exception as e:
            self.set_status(500)
            print(f"Errore durante aggiunta azienda: {str(e)}")
            self.write({"success": False, "error": str(e)})

class UpdateAziendaHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return
            
        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return
                
            # Controlla che l'ID sia presente
            if not data.get('id'):
                self.set_status(400)
                self.write({"success": False, "error": "ID azienda mancante"})
                return
            
            # Validate input
            validation_errors = validate_azienda_data(data)
            if validation_errors:
                self.set_status(400)
                self.write({"success": False, "errors": validation_errors})
                return
                
            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            
            # Verifichiamo prima se esiste la colonna stage
            cursor.execute("PRAGMA table_info(aziende)")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns]
            stage_exists = 'stage' in column_names
            
            # Se non esiste, aggiungiamola
            if not stage_exists:
                cursor.execute("ALTER TABLE aziende ADD COLUMN stage BOOLEAN DEFAULT 0")
                conn.commit()
            
            cursor.execute("""
                UPDATE aziende
                SET nome = ?,
                    via = ?,
                    cap = ?,
                    comune = ?,
                    nome_referente = ?,
                    email_referente = ?,
                    cellulare_referente = ?,
                    disponibilita_lavoro_estivo = ?,
                    osservazioni = ?,
                    stage = ?
                WHERE id = ?
            """, (
                data.get("nome", ""),
                data.get("via", ""),
                data.get("cap", ""),
                data.get("comune", ""),
                data.get("nome_referente", ""),
                data.get("email_referente", ""),
                data.get("cellulare_referente", ""),
                data.get("disponibilita_lavoro_estivo", False),
                data.get("osservazioni", ""),
                data.get("stage_pcto", False),  # Modificato per accettare il campo stage_pcto dal frontend
                data.get("id")
            ))
            
            conn.commit()
            conn.close()
            
            self.write({"success": True})
        except Exception as e:
            self.set_status(500)
            print(f"Errore durante aggiornamento azienda: {str(e)}")
            self.write({"success": False, "error": str(e)})
class GetAziendeStageHandler(tornado.web.RequestHandler):
    def get(self):
        try:
            conn = sqlite3.connect('db.sqlite3')
            c = conn.cursor()
            
            # Verifichiamo prima se esiste la colonna stage
            c.execute("PRAGMA table_info(aziende)")
            columns = c.fetchall()
            column_names = [column[1] for column in columns]
            stage_exists = 'stage' in column_names
            
            # Se non esiste, aggiungiamola
            if not stage_exists:
                c.execute("ALTER TABLE aziende ADD COLUMN stage BOOLEAN DEFAULT 0")
                conn.commit()
            
            # Query solo le aziende con stage = True
            c.execute("SELECT * FROM aziende WHERE stage = 1")
            rows = c.fetchall()
            
            # Get column names
            c.execute("PRAGMA table_info(aziende)")
            column_info = c.fetchall()
            column_names = [column[1] for column in column_info]
            
            # Create a list of dictionaries using column names
            result = []
            for row in rows:
                item = {}
                for i, value in enumerate(row):
                    if i < len(column_names):
                        item[column_names[i]] = value
                result.append(item)
            conn.close()
            self.write(json.dumps(result))
        except Exception as e:
            self.set_status(500)
            print(f"Errore durante recupero aziende con stage: {str(e)}")
            self.write(json.dumps({"error": str(e)}))

class DeleteAziendaHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return
            
        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return
                
            # Controlla che l'ID sia presente
            if not data.get('id'):
                self.set_status(400)
                self.write({"success": False, "error": "ID azienda mancante"})
                return
            
            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            
            # Prima eliminiamo tutte le associazioni legate all'azienda
            cursor.execute("DELETE FROM associazioni WHERE id_azienda = ?", (data.get('id'),))
            
            # Poi eliminiamo l'azienda
            cursor.execute("DELETE FROM aziende WHERE id = ?", (data.get('id'),))
            
            conn.commit()
            conn.close()
            
            self.write({"success": True})
        except Exception as e:
            self.set_status(500)
            print(f"Errore durante eliminazione azienda: {str(e)}")
            self.write({"success": False, "error": str(e)})
class MainHandler(tornado.web.RequestHandler):
    def get(self):
        try:
            show_login = self.get_argument("showLogin", "false")
            
            # Recupera le aziende con stage = True per mostrarle nella pagina index
            conn = sqlite3.connect('db.sqlite3')
            c = conn.cursor()
            
            # Verifichiamo prima se esiste la colonna stage
            c.execute("PRAGMA table_info(aziende)")
            columns = c.fetchall()
            column_names = [column[1] for column in columns]
            stage_exists = 'stage' in column_names
            
            # Se non esiste, aggiungiamola
            if not stage_exists:
                c.execute("ALTER TABLE aziende ADD COLUMN stage BOOLEAN DEFAULT 0")
                conn.commit()
            
            # Query solo le aziende con stage = True
            c.execute("SELECT * FROM aziende WHERE stage = 1")
            rows = c.fetchall()
            
            # Get column names
            c.execute("PRAGMA table_info(aziende)")
            column_info = c.fetchall()
            column_names = [column[1] for column in column_info]
            
            # Create a list of dictionaries using column names
            aziende_stage = []
            for row in rows:
                item = {}
                for i, value in enumerate(row):
                    if i < len(column_names):
                        item[column_names[i]] = value
                aziende_stage.append(item)
            conn.close()
            
            # Passa le aziende alla pagina index
            self.render("index.html", show_login=show_login, aziende_stage=aziende_stage)
        except Exception as e:
            print(f"Errore in MainHandler: {str(e)}")
            # Fallback in caso di errore
            self.render("index.html", show_login=show_login, aziende_stage=[])

class LoginPageHandler(tornado.web.RequestHandler):
    def get(self):
        self.render("login.html")

class LoginHandler(tornado.web.RequestHandler):
    def post(self):
        username = self.get_argument("username")
        password = self.get_argument("password")
        if username == USERNAME and password == PASSWORD:
            self.set_secure_cookie("user", username)
            self.redirect("/dashboard")
        else:
            self.write("Credenziali errate, riprova.")
class LogoutHandler(tornado.web.RequestHandler):
    def get(self):
        # Elimina il cookie di sessione
        self.clear_cookie("user")
        # Reindirizza alla homepage con parametro per mostrare login
        self.redirect("/?showLogin=true")

class DashboardHandler(tornado.web.RequestHandler):
    def get(self):
        user_cookie = self.get_secure_cookie("user")
        if not user_cookie:
            self.redirect("/login")
            return
        
        # Verifica che il cookie contenga effettivamente le credenziali corrette
        try:
            username = user_cookie.decode('utf-8')
            if username != USERNAME:
                self.clear_cookie("user")  # Cookie non valido, lo rimuoviamo
                self.redirect("/login")
                return
        except:
            self.clear_cookie("user")  # Cookie malformato, lo rimuoviamo
            self.redirect("/login")
            return
            
        self.render("dashboard.html")

class SecondHandler(tornado.web.RequestHandler):
    def get(self):
        # Renderizza il form iniziale (eventualmente con un messaggio di risultato)
        message = self.get_argument("message", "")
        self.render("sito_mail.html", message=message)

    def post(self):
        # Authentication check
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            
            return
        try:
            # Legge i dati usando i nomi definiti nel template in minuscolo
            anno_scolastico = self.get_argument("anno_scolastico")
            inizio_primo_periodo = self.get_argument("inizio_primo_periodo")
            fine_primo_periodo = self.get_argument("fine_primo_periodo")
            inizio_secondo_periodo = self.get_argument("inizio_secondo_periodo")
            fine_secondo_periodo = self.get_argument("fine_secondo_periodo")
            link_modulo = self.get_argument("link_modulo")
            prof_1 = self.get_argument("prof_1")
            prof_2 = self.get_argument("prof_2")
            print("Dati ricevuti")
        
            with open("mail.txt", "r", encoding="utf-8") as file:
                testo = file.read()
            testo = testo.replace("[[ANNO_SCOLASTICO]]", anno_scolastico)
            testo = testo.replace("[[INIZIO_PRIMO_PERIODO]]", inizio_primo_periodo)
            testo = testo.replace("[[FINE_PRIMO_PERIODO]]", fine_primo_periodo)
            testo = testo.replace("[[INIZIO_SECONDO_PERIODO]]", inizio_secondo_periodo)
            testo = testo.replace("[[FINE_SECONDO_PERIODO]]", fine_secondo_periodo)
            testo = testo.replace("[[LINK_MODULO]]", link_modulo)
            testo = testo.replace("[[PROF_1]]", prof_1)
            testo = testo.replace("[[PROF_2]]", prof_2)
            print("Mail modificata")
            
            # Renderizza la pagina visualizza_mail.html passando il testo della mail
            self.render("visualizza_mail.html", mail_text=testo, message="")
        except Exception as e:
            # Gestione degli errori
            error_msg = f"Errore durante la preparazione della mail: {str(e)}"
            print(error_msg)
            self.render("sito_mail.html", message=error_msg)


class PreviewEmailHandler(tornado.web.RequestHandler):
    def post(self):
        # Questo handler gestirà solo le richieste POST al form della pagina di anteprima
        try:
            mail_text = self.get_argument("mail_text")
            self.render("visualizza_mail.html", mail_text=mail_text, message="")
        except Exception as e:
            print(f"Errore durante l'anteprima della mail: {str(e)}")
            self.redirect("/sito_mail?message=" + tornado.escape.url_escape(f"Errore: {str(e)}"))


class SendEmailHandler(tornado.web.RequestHandler):
    def post(self):
        # Authentication check
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return
        try:
            mail_text = self.get_argument("mail_text")
            subject = "Richiesta disponibilità PCTO"

            # Connessione al database
            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            cursor.execute("SELECT email_referente FROM aziende WHERE email_referente != ''")
            results = cursor.fetchall()
            conn.close()
            
            # Estrae gli indirizzi email dalla query
            emails = [row[0] for row in results if row[0]]
            print(f"Invio della mail a {len(emails)} indirizzi.")
            
            # Contatori per monitorare il risultato dell'invio
            success_count = 0
            error_count = 0
            error_messages = []

            # Itera su ogni email e invia la mail
            for email in emails:
                try:
                    mail_handler.send_email("alessandra.mogliazza@fermi.mo.it", 
                                           settings.GOOGLE_APP_KEY, 
                                           email, 
                                           subject, 
                                           mail_text)
                    success_count += 1
                except Exception as e:
                    error_message = f"Errore nell'invio a {email}: {e}"
                    print(error_message)
                    error_messages.append(error_message)
                    error_count += 1
            
            # Prepara il messaggio di risultato
            result_message = f"Email inviate con successo: {success_count}"
            if error_count > 0:
                result_message += f". Email non inviate: {error_count}"
            
            # Renderizza la pagina con messaggio di conferma
            self.redirect("/sito_mail?message=" + tornado.escape.url_escape(result_message))
            
        except Exception as e:
            error_msg = f"Errore generale nell'invio delle email: {str(e)}"
            print(error_msg)
            self.redirect("/sito_mail?message=" + tornado.escape.url_escape(error_msg))

# Add these imports at the top if not already present
import csv
import io
def validate_email(email):
    import re
    return re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email) is not None
# Add these validation functions if not already present in app.py
# Function to validate the class format (should be a number 4 followed by an uppercase letter)
def validate_classe(classe):
    import re
    return re.match(r'^[4][A-Z]$', classe) is not None

# Function to validate email (already exists in your code, so no need to duplicate)

# Add this handler class to your app.py file
class ImportStudentiCsvHandler(tornado.web.RequestHandler):
    def post(self):
        # Authentication check
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return
            
        try:
            # Get CSV content from the request
            csv_data = self.request.body.decode('utf-8')
            if not csv_data:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato CSV ricevuto"})
                return
                
            # Create database connection
            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            
            # Check if the students table has a class field
            cursor.execute("PRAGMA table_info(studenti)")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns]
            classe_exists = 'classe' in column_names
            
            # If the class field doesn't exist, add it
            if not classe_exists:
                cursor.execute("ALTER TABLE studenti ADD COLUMN classe TEXT")
                conn.commit()
                classe_exists = True
                
            # Process the CSV
            csv_file = io.StringIO(csv_data)
            csv_reader = csv.DictReader(csv_file)
            
            # Counters for statistics
            inserted = 0
            updated = 0
            errors = []
            
            # Process each row of the CSV
            for row_num, row in enumerate(csv_reader, start=2):  # Start=2 because row 1 is the header
                try:
                    # Check that all necessary fields are present
                    if not all(key in row and row[key].strip() for key in ['nome', 'email', 'classe']):
                        errors.append(f"Riga {row_num}: Campi mancanti o vuoti (nome, email, classe richiesti)")
                        continue
                        
                    # Validate class format
                    if not validate_classe(row['classe']):
                        errors.append(f"Riga {row_num}: Il formato della classe non è valido '{row['classe']}' (deve essere un numero (4) seguito da una lettera maiuscola)")
                        continue
                        
                    # Validate email format
                    if not validate_email(row['email']):
                        errors.append(f"Riga {row_num}: Email non valida '{row['email']}'")
                        continue
                        
                    # Check if a student with this email already exists
                    cursor.execute("SELECT id FROM studenti WHERE email = ?", (row['email'],))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # Update existing student
                        cursor.execute("""
                            UPDATE studenti
                            SET nome = ?, classe = ?
                            WHERE id = ?
                        """, (row['nome'], row['classe'], existing[0]))
                        updated += 1
                    else:
                        # Insert new student
                        cursor.execute("""
                            INSERT INTO studenti (nome, email, classe)
                            VALUES (?, ?, ?)
                        """, (row['nome'], row['email'], row['classe']))
                        inserted += 1
                        
                except Exception as e:
                    errors.append(f"Riga {row_num}: Errore - {str(e)}")
            
            # Commit changes to the database
            conn.commit()
            conn.close()
            
            # Return import results
            self.write({
                "success": True,
                "inserted": inserted,
                "updated": updated,
                "errors": errors
            })
            
        except Exception as e:
            self.set_status(500)
            print(f"Errore durante l'importazione CSV: {str(e)}")
            self.write({"success": False, "error": str(e)})

class StudentiHandler(tornado.web.RequestHandler):
    def get(self):
        conn = sqlite3.connect("db.sqlite3")
        cursor = conn.cursor()
        
        # Prima verifichiamo la struttura della tabella
        cursor.execute("PRAGMA table_info(studenti)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns]
        
        # Verifichiamo che il campo 'classe' esista
        classe_exists = 'classe' in column_names
        
        # Query con gestione appropriata dei campi
        cursor.execute("SELECT * FROM studenti")
        rows = cursor.fetchall()
        
        # Costruiamo un risultato con sicurezza per i campi mancanti
        result = []
        for r in rows:
            student = {
                "id": r[0],
                "nome": r[1] if len(r) > 1 else "",
                "email": r[2] if len(r) > 2 else ""
            
            }
            
            # Aggiungi classe solo se esiste nella tabella
            if classe_exists and len(r) > 3:
                student["classe"] = r[3] if r[3] is not None else ""
            else:
                student["classe"] = ""
                
            result.append(student)
            
        conn.close()
        self.write(json.dumps(result))

class AddStudenteHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return

        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            # Debug: stampa il corpo della richiesta
            print("Request body:", self.request.body)
            
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return

            # Debug: stampa i dati parsati
            print("Parsed data:", data)
            
            # Assicurati che data sia un dizionario
            if not isinstance(data, dict):
                self.set_status(400)
                self.write({"success": False, "error": "Il formato dei dati non è valido. Atteso un oggetto JSON."})
                return

            # Validate input
            validation_errors = validate_student_data(data)
            if validation_errors:
                self.set_status(400)
                self.write({"success": False, "errors": validation_errors})
                return

            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            
            # Verifichiamo se il campo classe esiste nella tabella
            cursor.execute("PRAGMA table_info(studenti)")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns]
            
            if 'classe' in column_names:
                # Se esiste, includiamo il campo nella query
                cursor.execute("""
                    INSERT INTO studenti (nome, email, classe)
                    VALUES (?, ?, ?)
                """, (data.get('nome', ''), data.get('email', ''), data.get('classe', '')))
            else:
                # Altrimenti inseriamo solo nome ed email
                cursor.execute("""
                    INSERT INTO studenti (nome, email)
                    VALUES (?, ?)
                """, (data.get('nome', ''), data.get('email', '')))

            last_id = cursor.lastrowid
            conn.commit()
            conn.close()

            self.write({"success": True, "id": last_id})

        except Exception as e:
            self.set_status(500)
            print(f"Errore durante aggiunta studente: {str(e)}")
            self.write({"success": False, "error": str(e)})

class UpdateStudenteHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.write({"success": False, "error": "Non autorizzato"})
            return

        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return
                
            # Assicurati che data sia un dizionario
            if not isinstance(data, dict):
                self.set_status(400)
                self.write({"success": False, "error": "Il formato dei dati non è valido. Atteso un oggetto JSON."})
                return
                
            # Controlla che l'ID sia presente
            if not data.get('id'):
                self.set_status(400)
                self.write({"success": False, "error": "ID studente mancante"})
                return

            # Validate input
            validation_errors = validate_student_data(data)
            if validation_errors:
                self.set_status(400)
                self.write({"success": False, "errors": validation_errors})
                return

            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            
            # Verifichiamo se il campo classe esiste nella tabella
            cursor.execute("PRAGMA table_info(studenti)")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns]
            
            if 'classe' in column_names:
                # Se esiste, includiamo il campo nella query
                cursor.execute("""
                    UPDATE studenti
                    SET nome = ?, email = ?, classe = ?
                    WHERE id = ?
                """, (data.get('nome', ''), data.get('email', ''), data.get('classe', ''), data.get('id')))
            else:
                # Altrimenti aggiorniamo solo nome ed email
                cursor.execute("""
                    UPDATE studenti
                    SET nome = ?, email = ?
                    WHERE id = ?
                """, (data.get('nome', ''), data.get('email', ''), data.get('id')))

            conn.commit()
            conn.close()

            self.write({"success": True})

        except Exception as e:
            self.set_status(500)
            print(f"Errore durante aggiornamento studente: {str(e)}")
            self.write({"success": False, "error": str(e)})

class DeleteStudenteHandler(tornado.web.RequestHandler):
    def post(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.render("login.html")
            self.write({"success": False, "error": "Non autorizzato"})
            return

        try:
            # Gestione errore: assicurati che il corpo della richiesta esista
            if not self.request.body:
                self.set_status(400)
                self.write({"success": False, "error": "Nessun dato ricevuto"})
                return
                
            try:
                data = json.loads(self.request.body)
            except json.JSONDecodeError as e:
                self.set_status(400)
                self.write({"success": False, "error": f"JSON non valido: {str(e)}"})
                return
                
            # Controlla che l'ID sia presente
            if not data.get('id'):
                self.set_status(400)
                self.write({"success": False, "error": "ID studente mancante"})
                return

            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()

            # First, remove associated associations
            cursor.execute("DELETE FROM associazioni WHERE id_studente = ?", (data.get('id'),))
            # Then delete the student
            cursor.execute("DELETE FROM studenti WHERE id = ?", (data.get('id'),))

            conn.commit()
            conn.close()

            self.write({"success": True})

        except Exception as e:
            self.set_status(500)
            print(f"Errore durante eliminazione studente: {str(e)}")
            self.write({"success": False, "error": str(e)})
            
import io
import csv
import datetime
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
class ExportAssociazioniHandler(tornado.web.RequestHandler):
    def get(self):
        # Authentication check
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.render("login.html")
            self.write({"success": False, "error": "Non autorizzato"})
            return
        # Check if an export file was requested
        if self.get_argument("format", None):
            # Export data request
            self.export_data()
        else:
            # Load the HTML page
            try:
                with open("excel-export.html", "r", encoding="utf-8") as file:
                    html_content = file.read()
                
                # Send HTML content as response
                self.write(html_content)
            except Exception as e:
                error_msg = f"Error loading export page: {str(e)}"
                print(error_msg)
                self.set_status(500)
                self.write(error_msg)
    
    def export_data(self):
        try:
            # Get the requested format (default: csv)
            export_format = self.get_argument("format", "csv")
            
            # Connect to database
            conn = sqlite3.connect("db.sqlite3")
            cursor = conn.cursor()
            
            # Check if the "classe" field exists in the "studenti" table
            cursor.execute("PRAGMA table_info(studenti)")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns]
            classe_exists = 'classe' in column_names
            
            # Execute query to get association data
            if classe_exists:
                cursor.execute("""
                    SELECT 
                        a.id, 
                        s.nome as studente_nome, 
                        s.email as studente_email,
                        s.classe as studente_classe,
                        az.nome as azienda_nome, 
                        az.via as azienda_via,
                        az.cap as azienda_cap,
                        az.comune as azienda_comune,
                        az.nome_referente,
                        az.email_referente,
                        az.cellulare_referente,
                        a.data_inizio, 
                        a.data_fine
                    FROM associazioni a
                    JOIN studenti s ON a.id_studente = s.id
                    JOIN aziende az ON a.id_azienda = az.id
                    ORDER BY s.nome
                """)
            else:
                cursor.execute("""
                    SELECT 
                        a.id, 
                        s.nome as studente_nome, 
                        s.email as studente_email,
                        '' as studente_classe,
                        az.nome as azienda_nome, 
                        az.via as azienda_via,
                        az.cap as azienda_cap,
                        az.comune as azienda_comune,
                        az.nome_referente,
                        az.email_referente,
                        az.cellulare_referente,
                        a.data_inizio, 
                        a.data_fine
                    FROM associazioni a
                    JOIN studenti s ON a.id_studente = s.id
                    JOIN aziende az ON a.id_azienda = az.id
                    ORDER BY s.nome
                """)
            
            rows = cursor.fetchall()
            
            # Define column headers
            headers = [
                "ID", "Nome Studente", "Email Studente", "Classe",
                "Nome Azienda", "Via", "CAP", "Comune", 
                "Nome Referente", "Email Referente", "Cellulare Referente",
                "Data Inizio", "Data Fine"
            ]
            
            # Current timestamp for filename
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if export_format == "csv":
                # Create an in-memory buffer for CSV
                output = io.StringIO()
                writer = csv.writer(output)
                
                # Write header
                writer.writerow(headers)
                
                # Write data
                for row in rows:
                    writer.writerow(row)
                
                # Prepare to send the file
                self.set_header("Content-Type", "text/csv")
                self.set_header("Content-Disposition", f"attachment; filename=associazioni_{timestamp}.csv")
                
                self.write(output.getvalue())
                
            elif export_format == "excel":
                # Create a workbook and select the active sheet
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Associazioni"
                
                # Header style
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Add header
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Add data
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                
                # Save the file to an in-memory buffer
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Prepare to send the file
                self.set_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.set_header("Content-Disposition", f"attachment; filename=associazioni_{timestamp}.xlsx")
                
                self.write(output.read())
                
            else:
                # If an unsupported format is requested, return an error
                self.set_status(400)
                self.write({"success": False, "error": f"Format '{export_format}' not supported"})
                
            conn.close()
            
        except Exception as e:
            self.set_status(500)
            error_msg = f"Error during export: {str(e)}"
            print(error_msg)
            self.write({"success": False, "error": error_msg})
# Handler per la gestione delle aziende
class GetAziendeHandler(tornado.web.RequestHandler):
    def get(self):
        try:
            conn = sqlite3.connect('db.sqlite3')
            c = conn.cursor()
            # Query the aziende table
            c.execute("SELECT * FROM aziende")
            rows = c.fetchall()
            # Get column names
            c.execute("PRAGMA table_info(aziende)")
            column_names = [column[1] for column in c.fetchall()]
            # Create a list of dictionaries using column names
            result = []
            for row in rows:
                item = {}
                for i, value in enumerate(row):
                    if i < len(column_names):
                        item[column_names[i]] = value
                result.append(item)
            conn.close()
            self.write(json.dumps(result))
        except Exception as e:
            self.set_status(500)
            print(f"Errore durante recupero aziende: {str(e)}")
            self.write(json.dumps({"error": str(e)}))
import os
import sqlite3
import tornado.ioloop
import tornado.web
from convenzioni_handler import replace_text_in_docx

DB_FILE = 'db.sqlite3'

def fetch_data_from_db(associazione_id):
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            studenti.nome AS STUDENTE,
            aziende.nome AS AZIENDA,
            aziende.via AS VIA,
            aziende.comune AS CITTA_SEDE,
            aziende.nome_referente AS NOME_RAP,
            associazioni.data_inizio AS INIZIO_PERIODO,
            associazioni.data_fine AS FINE_PERIODO
        FROM associazioni
        JOIN studenti ON associazioni.id_studente = studenti.id
        JOIN aziende ON associazioni.id_azienda = aziende.id
        WHERE associazioni.id = ?;
    """, (associazione_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        raise tornado.web.HTTPError(404, f"Nessuna convenzione trovata per id={associazione_id}")
    return dict(row)

def get_all_associazioni():
    """Recupera tutti gli ID delle associazioni dal database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM associazioni")
    associazioni_ids = [row[0] for row in cursor.fetchall()]
    conn.close()
    return associazioni_ids

def process_all_convenzioni():
    """Elabora tutte le convenzioni per ogni associazione nel database"""
    template_file = 'template.docx'
    output_dir = 'convenzioni_generate'
    
    # Crea la directory di output se non esiste già
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Ottieni tutti gli ID delle associazioni
    associazioni_ids = get_all_associazioni()
    
    for associazione_id in associazioni_ids:
        try:
            # Preleva i dati per questa associazione
            data = fetch_data_from_db(associazione_id)
            
            # Genera il nome del file di output
            output_file = os.path.join(output_dir, f"{data['AZIENDA']}_convenzione.docx")
            
            # Usa la funzione fornita per sostituire i segnaposto nel template
            replace_text_in_docx(
                template_file,
                output_file,
                {
                    "[[STUDENTE]]": data['STUDENTE'],
                    "[[AZIENDA]]": data['AZIENDA'],
                    "[[VIA]]": data['VIA'],
                    "[[CITTA_SEDE]]": data['CITTA_SEDE'],
                    "[[NOME_RAP]]": data['NOME_RAP'],
                    "[[INIZIO_PERIODO]]": data['INIZIO_PERIODO'],
                    "[[FINE_PERIODO]]": data['FINE_PERIODO']
                }
            )
            
            print(f"Generata convenzione per associazione ID {associazione_id}: {output_file}")
            
        except Exception as e:
            print(f"Errore nella generazione della convenzione per associazione ID {associazione_id}: {str(e)}")

# --- HANDLER HTTP ---
class ConvenzioneHandler(tornado.web.RequestHandler):

    async def get(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.render("login.html")
            self.write({"success": False, "error": "Non autorizzato"})
            return
        action = self.get_argument('action', 'single')
        
        if action == 'all':
            # Elabora tutte le convenzioni
            process_all_convenzioni()
            self.write("Tutte le convenzioni sono state generate nella directory 'convenzioni_generate'")
            return
        
        # Funzionalità originale per una singola convenzione
        convenzione_id = self.get_argument('convenzione_id', None)
        if not convenzione_id:
            raise tornado.web.HTTPError(400, "Parametro 'convenzione_id' obbligatorio quando action non è 'all'")

        try:
            # Preleva dati
            data = fetch_data_from_db(convenzione_id)
            
            # Usa la funzione fornita in convenzioni_handler.py
            template_file = 'template.docx'
            output_file = f"Convenzione_{convenzione_id}.docx"
            
            replacements = {
                
                "[[STUDENTE]]": data['STUDENTE'],
                "[[AZIENDA]]": data['AZIENDA'],
                "[[VIA]]": data['VIA'],
                "[[CITTA_SEDE]]": data['CITTA_SEDE'],
                "[[NOME_RAP]]": data['NOME_RAP'],
                "[[INIZIO_PERIODO]]": data['INIZIO_PERIODO'],
                "[[FINE_PERIODO]]": data['FINE_PERIODO']
            }
            
            replace_text_in_docx(template_file, output_file, replacements)
            
            # Invia il file come download
            self.set_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            self.set_header('Content-Disposition', f'attachment; filename="{output_file}"')
            with open(output_file, 'rb') as f:
                self.write(f.read())
            
            # Rimuovi il file temporaneo dopo l'invio
            if os.path.exists(output_file):
                os.remove(output_file)
                
        except tornado.web.HTTPError:
            raise
        except Exception as e:
            raise tornado.web.HTTPError(500, f"Errore durante la generazione della convenzione: {str(e)}")

class ProcessAllHandler(tornado.web.RequestHandler):
    def get(self):
        if not self.get_secure_cookie("user"):
            self.set_status(403)
            self.render("login.html")
            self.write({"success": False, "error": "Non autorizzato"})
            return
        
        # This gets called when the user confirms via the popup
        process_all_convenzioni()
        self.render("convenzioniii.html")
        self.write({"success": True, "message": "Elaborazione di tutte le convenzioni completata. I file sono stati salvati nella directory 'convenzioni_generate'"})

class ProcessOneHandler(tornado.web.RequestHandler):
  def get(self):
        self.render("convezioni.html")
def make_app():
    return tornado.web.Application([
       (r"/", MainHandler),
        (r"/login", LoginPageHandler),
        (r"/auth", LoginHandler),
        (r"/logout", LogoutHandler),
        (r"/dashboard", DashboardHandler),

        # Student Management Endpoints
        (r"/get-studenti", StudentiHandler),
        (r"/add-studente", AddStudenteHandler),
        (r"/update-studente", UpdateStudenteHandler),
        (r"/delete-studente", DeleteStudenteHandler),
        (r"/import-studenti-csv", ImportStudentiCsvHandler),

        # Association Management Endpoints
        (r"/get-associazioni", GetAssociazioniHandler),
        (r"/add-associazione", AddAssociazioneHandler),
        (r"/update-associazione", UpdateAssociazioneHandler),
        (r"/delete-associazione", DeleteAssociazioneHandler),

        # Company Management Endpoints
        (r"/get-aziende", GetAziendeHandler),
        (r"/get-aziende-stage", GetAziendeStageHandler),  # Nuovo endpoint per aziende con stage
        (r"/add-azienda", AddAziendaHandler),
        (r"/update-azienda", UpdateAziendaHandler),
        (r"/delete-azienda", DeleteAziendaHandler),
        (r"/sito_mail", SecondHandler),
        (r"/preview_email", PreviewEmailHandler),  
        (r"/send_email", SendEmailHandler),
        (r"/export-associazioni", ExportAssociazioniHandler),
        (r"/generate", ConvenzioneHandler),
        (r"/process-all", ProcessAllHandler),
        (r"/convenzioni", ProcessOneHandler),



        (r"/static/(.*)", tornado.web.StaticFileHandler, {"path": "static"}),
    ],
    template_path="templates",
    static_path="static",
    cookie_secret="my_secret_key",
    login_url="/login",
    xsrf_cookies=False
    )

if __name__ == "__main__":
    app = make_app()
    app.listen(8100)
    print("Server running on http://localhost:8100")
    tornado.ioloop.IOLoop.current().start()
